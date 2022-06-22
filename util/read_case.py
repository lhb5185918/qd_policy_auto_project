import xlrd
import openpyxl
import os
from config.path_data import log_file_path
from config.path_data import case_path
from openpyxl import load_workbook

class Case_operation:
    def get_case(self,filename,sheetname):
        wb = xlrd.open_workbook(filename=filename)
        sheet = wb.sheet_by_name(sheetname)
        list1 = []
        case = []
        key = sheet.row_values(0)
        for i in range(1,sheet.nrows):
            d = []
            res = sheet.row_values(i)
            for a in res :
                d.append(a)
            list1.append(d)
            for s in list1:
                r = dict(zip(key,s))
            case.append(r)
        return case

    def write_result(self,res,xulie):#测试结果写入
        wb = open(file=log_file_path, mode='r', encoding='utf-8').readlines()#打开log文件
        wc = load_workbook(case_path)#打开excle文件
        data = []
        l = []
        for i in wb:
            data.append(i)#将文本文件的值存储至列表中
        for a in data:
            if res in a:
                ws = wc.active
                sheet = wc.worksheets[xulie]
                l.append(a)
        for i in range(len(l)):
            sheet.cell(i + 2, 7).value = l[i]
        wc.save(case_path)


