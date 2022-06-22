import json
from config.path_data import case_path
import pytest
import requests
from util.request_util import Requests
from util.mysql_util import Pysql_util
from config.path_data import case_path
import time

import  openpyxl
from openpyxl import Workbook,load_workbook

wb = load_workbook(case_path)
ws = wb.active
sheet = wb.worksheets[0]
l = [1,2,3]
for i in range(len(l)):
    sheet.cell(i + 2, 7).value = l[i]
wb.save(case_path)

a = "123"
print(a[0:2])






