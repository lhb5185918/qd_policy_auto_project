from config.path_data import log_file_path
from util.read_case import Case_operation
from config.path_data import case_path
from openpyxl import load_workbook
wb =open(file=log_file_path,mode='r',encoding='utf-8').readlines()
wc = load_workbook(case_path)
count = len(wb)
data =[]
l=[]
for i in wb:
    data.append(i)
for a in data :
    if '新增政策' in a:
        ws = wc.active
        sheet = wc.worksheets[2]
        l.append(a)
print(l)
for i in range(len(l)):
    sheet.cell(i + 2, 7).value = l[i]
wc.save(case_path)







